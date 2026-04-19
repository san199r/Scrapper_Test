import os
import re
import time
from datetime import datetime, timedelta

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


SEARCH_URL = "https://aca-prod.accela.com/anaheim/Cap/CapHome.aspx?module=Building&TabName=Building"
DATE_INPUT_FILE = os.path.join(os.getcwd(), "Date_Input.xlsx")
OUTPUT_FILE = os.path.join(
    os.getcwd(),
    f"Anaheim_CA_Permits_DateRange_Tool_Output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

HEADERS = [
    "S.No",
    "County",
    "State",
    "Record Number",
    "Date",
    "Record Type",
    "Project Name",
    "Address",
    "Status",
    "Action",
    "Description",
    "Expiration Date",
    "Short Notes",
    "Applicant Name",
    "Applicant Company Name",
    "Applicant Address",
    "Applicant Mailing Address",
    "Applicant Phone No",
    "Applicant Mail",
    "Applicant Work Phone",
    "Licensed Professional Name",
    "Licensed Professional Mail",
    "Licensed Professional Company Name",
    "Licensed Professional Address",
    "Licensed Professional Mailing City, State, Zip",
    "Licensed Professional Business Phone",
    "Licensed Professional Tax No",
    "Licensed Professional Contractor No",
    "Project Description",
    "Owner",
    "Owner Address",
    "Owner Mailing City, State, Zip",
    "Parcel No",
]


def log_message(message):
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {message}", flush=True)


def clean_text(value):
    if value is None:
        return ""
    value = str(value).replace("\xa0", " ")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\s*\n\s*", "\n", value)
    return value.strip()


def create_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument(
        "--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_experimental_option(
        "excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.set_page_load_timeout(120)
    return driver


def wait_ready(driver, timeout=30):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )


def init_excel(path):
    if os.path.exists(path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Anaheim Records"
    ws.append(HEADERS)
    wb.save(path)


def append_row_to_excel(path, row_data):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Anaheim Records"
        ws.append(HEADERS)

    ws.append([row_data.get(col, "") for col in HEADERS])
    wb.save(path)


def get_existing_record_numbers(path):
    existing = set()
    if not os.path.exists(path):
        return existing

    wb = load_workbook(path, read_only=True)
    ws = wb.active

    header_row = next(
        ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True),
        None)
    if not header_row:
        wb.close()
        return existing

    try:
        rec_idx = list(header_row).index("Record Number")
    except ValueError:
        wb.close()
        return existing

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > rec_idx and row[rec_idx]:
            existing.add(clean_text(row[rec_idx]))

    wb.close()
    return existing


def get_next_serial_no(path):
    if not os.path.exists(path):
        return 1

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    max_row = ws.max_row
    wb.close()

    if max_row <= 1:
        return 1
    return max_row


def get_date_range(path=None):
    """Returns (from_date, to_date). Defaults to last 30 days if no path or file missing."""
    today = datetime.now()
    start = today - timedelta(days=30)

    default_from = start.strftime("%m/%d/%Y")
    default_to = today.strftime("%m/%d/%Y")

    if not path or not os.path.exists(path):
        log_message(
            f"No date input file specified or found ({path}). Defaulting to last 30 days: {default_from} to {default_to}")
        return default_from, default_to

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        from_date = ws["A2"].value
        to_date = ws["B2"].value
        wb.close()

        if from_date and to_date:
            if isinstance(from_date, datetime):
                from_date = from_date.strftime("%m/%d/%Y")
            if isinstance(to_date, datetime):
                to_date = to_date.strftime("%m/%d/%Y")
            return str(from_date), str(to_date)
    except Exception as e:
        log_message(
            f"Error reading dates from Excel ({e}). Using default 30-day range.")

    return default_from, default_to


def perform_automated_search(driver, wait, start_date, end_date):
    log_message(f"Navigating to Building Search: {SEARCH_URL}")
    driver.get(SEARCH_URL)
    wait_ready(driver, 30)
    time.sleep(2)

    log_message("Waiting for date fields...")
    in_iframe = False
    try:
        # standard wait
        start_date_field = wait.until(EC.visibility_of_element_located(
            (By.ID, "ctl00_PlaceHolderMain_generalSearchForm_txtGSStartDate")))
        end_date_field = wait.until(EC.visibility_of_element_located(
            (By.ID, "ctl00_PlaceHolderMain_generalSearchForm_txtGSEndDate")))
    except Exception:
        log_message(
            "Date fields not found with standard IDs. Checking for iframes...")
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for i, iframe in enumerate(iframes):
            driver.switch_to.frame(iframe)
            try:
                start_date_field = driver.find_element(
                    By.ID, "ctl00_PlaceHolderMain_generalSearchForm_txtGSStartDate")
                end_date_field = driver.find_element(
                    By.ID, "ctl00_PlaceHolderMain_generalSearchForm_txtGSEndDate")
                log_message(f"Found date fields in iframe {i}")
                in_iframe = True
                break
            except Exception:
                driver.switch_to.default_content()

    if not in_iframe:
        # Re-fetch just in case
        start_date_field = driver.find_element(
            By.ID, "ctl00_PlaceHolderMain_generalSearchForm_txtGSStartDate")
        end_date_field = driver.find_element(
            By.ID, "ctl00_PlaceHolderMain_generalSearchForm_txtGSEndDate")

    # Hard-set values via JS to bypass masked input issues
    for field, val in [(start_date_field, start_date),
                       (end_date_field, end_date)]:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", field)
        driver.execute_script("arguments[0].value = arguments[1];", field, val)
        driver.execute_script(
            "arguments[0].dispatchEvent(new Event('change'));", field)
        driver.execute_script(
            "arguments[0].dispatchEvent(new Event('blur'));", field)

    time.sleep(2)

    log_message("Clicking Search Button...")
    search_btn = driver.find_element(
        By.ID, "ctl00_PlaceHolderMain_btnNewSearch")
    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});",
        search_btn)
    time.sleep(1)

    try:
        search_btn.click()
    except Exception:
        driver.execute_script("arguments[0].click();", search_btn)

    log_message("Waiting for results...")
    wait_ready(driver, 30)
    time.sleep(5)


def ensure_results_frame(driver):
    """Ensures the driver is switched to the iframe containing results, or default content if no iframe."""
    driver.switch_to.default_content()
    # First check if the table is already in default content
    try:
        if driver.find_elements(
                By.ID,
                "ctl00_PlaceHolderMain_dgvPermitList_gdvPermitList"):
            return True
    except BaseException:
        pass

    # Otherwise check iframes
    iframes = driver.find_elements(By.TAG_NAME, "iframe")
    for iframe in iframes:
        try:
            driver.switch_to.frame(iframe)
            if driver.find_elements(
                    By.ID, "ctl00_PlaceHolderMain_dgvPermitList_gdvPermitList"):
                return True
        except BaseException:
            pass
        driver.switch_to.default_content()

    # Fallback to search fields iframe if results table not found yet (maybe
    # search was empty)
    for iframe in iframes:
        try:
            driver.switch_to.frame(iframe)
            if driver.find_elements(
                    By.ID, "ctl00_PlaceHolderMain_generalSearchForm_txtGSStartDate"):
                return True
        except BaseException:
            pass
        driver.switch_to.default_content()

    return False


def get_result_table(driver, wait):
    ensure_results_frame(driver)
    return wait.until(EC.presence_of_element_located(
        (By.XPATH, '//*[@id="ctl00_PlaceHolderMain_dgvPermitList_gdvPermitList"]')))


def detect_header_map(table):
    header_map = {}
    rows = table.find_elements(By.XPATH, "./tbody/tr")

    for row in rows:
        ths = row.find_elements(By.TAG_NAME, "th")
        if ths:
            for i, th in enumerate(ths):
                txt = clean_text(th.text).lower()
                if txt:
                    header_map[txt] = i
            if header_map:
                return header_map

    # Anaheim specific fallback
    fallback = [
        "",
        "date",
        "record number",
        "record type",
        "address",
        "status",
        "description",
        "expiration date",
        "action",
    ]
    for i, name in enumerate(fallback):
        if name:
            header_map[name] = i
    return header_map


def get_all_data_rows_count(driver):
    ensure_results_frame(driver)
    rows = driver.find_elements(
        By.XPATH,
        '//*[@id="ctl00_PlaceHolderMain_dgvPermitList_gdvPermitList"]/tbody/tr[td]')
    count = 0
    for row in rows:
        try:
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds) < 5:
                continue
            if row.find_elements(
                    By.XPATH, './/table[contains(@class,"aca_pagination")]'):
                continue
            first_cell_text = clean_text(tds[0].text) if len(tds) > 0 else ""
            second_cell_text = clean_text(tds[1].text) if len(tds) > 1 else ""
            if "record number" in first_cell_text.lower() or "date" in first_cell_text.lower():
                continue
            if second_cell_text:
                count += 1
        except Exception:
            pass
    return count


def get_data_row_by_position(driver, position):
    ensure_results_frame(driver)
    rows = driver.find_elements(
        By.XPATH,
        '//*[@id="ctl00_PlaceHolderMain_dgvPermitList_gdvPermitList"]/tbody/tr[td]')

    valid_rows = []
    for row in rows:
        try:
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds) < 5:
                continue
            if row.find_elements(
                    By.XPATH, './/table[contains(@class,"aca_pagination")]'):
                continue
            first_cell_text = clean_text(tds[0].text) if len(tds) > 0 else ""
            second_cell_text = clean_text(tds[1].text) if len(tds) > 1 else ""
            if "record number" in first_cell_text.lower() or "date" in first_cell_text.lower():
                continue
            if second_cell_text:
                valid_rows.append(row)
        except Exception:
            pass

    if position < 1 or position > len(valid_rows):
        raise IndexError(
            f"Row position {position} out of range. Found {len(valid_rows)} valid rows.")

    return valid_rows[position - 1]


def extract_row_cells(row):
    cells = row.find_elements(By.TAG_NAME, "td")
    return [clean_text(td.text) for td in cells]


def pick_value(cells, header_map, *possible_names):
    for name in possible_names:
        idx = header_map.get(name.lower())
        if idx is not None and idx < len(cells):
            return cells[idx]
    return ""


def parse_summary_from_row(driver, row, wait):
    table = get_result_table(driver, wait)
    header_map = detect_header_map(table)
    cells = extract_row_cells(row)

    summary = {
        "Record Number": pick_value(cells, header_map, "record number"),
        "Date": pick_value(cells, header_map, "date", "application date"),
        "Record Type": pick_value(cells, header_map, "record type"),
        "Address": pick_value(cells, header_map, "address"),
        "Status": pick_value(cells, header_map, "status"),
        "Description": pick_value(cells, header_map, "description"),
        "Expiration Date": pick_value(cells, header_map, "expiration date"),
        "Action": pick_value(cells, header_map, "action"),
    }

    # Fallback if pick_value failed
    if not summary["Record Number"] and len(cells) > 2:
        summary["Record Number"] = cells[2]
    if not summary["Date"] and len(cells) > 1:
        summary["Date"] = cells[1]

    return summary


def is_record_clickable(row):
    try:
        links = row.find_elements(
            By.XPATH, './/a[contains(@id,"PermitNumber")]')
        if not links:
            return False
        link = links[0]
        text = clean_text(link.text)
        if not text or not link.is_displayed():
            return False
        return True
    except Exception:
        return False


def open_record_in_new_tab(driver, row):
    link = row.find_element(By.XPATH, './/a[contains(@id,"PermitNumber")]')
    href = (link.get_attribute("href") or "").strip()
    old_handles = driver.window_handles

    if href and not href.lower().startswith("javascript:void(0)"):
        driver.execute_script("window.open(arguments[0], '_blank');", href)
    else:
        driver.execute_script("arguments[0].click();", link)

    WebDriverWait(
        driver, 20).until(
        lambda d: len(
            d.window_handles) > len(old_handles))
    new_handle = [h for h in driver.window_handles if h not in old_handles][0]
    driver.switch_to.window(new_handle)
    wait_ready(driver, 30)


def find_label_block(soup, label_text):
    pattern = re.compile(rf"^\s*{re.escape(label_text)}\s*$", re.I)
    label = soup.find("span", string=pattern)
    if not label:
        return None
    parent_h1 = label.find_parent("h1")
    if not parent_h1:
        return label.find_next("span")
    return parent_h1.find_next_sibling("span")


def extract_applicant(soup):
    result = {
        "Applicant Name": "",
        "Applicant Company Name": "",
        "Applicant Address": "",
        "Applicant Mailing Address": "",
        "Applicant Phone No": "",
        "Applicant Mail": "",
        "Applicant Work Phone": "",
    }

    block = find_label_block(soup, "Applicant:")
    if not block:
        return result

    business = block.select_one("span.contactinfo_businessname")
    if business:
        result["Applicant Company Name"] = clean_text(business.get_text())

    first = block.select_one("span.contactinfo_firstname")
    last = block.select_one("span.contactinfo_lastname")
    full_name = clean_text(
        f"{first.get_text() if first else ''} {last.get_text() if last else ''}")
    result["Applicant Name"] = full_name

    phone1 = block.select_one("span.contactinfo_phone1 .ACA_PhoneNumberLTR")
    if phone1:
        result["Applicant Phone No"] = clean_text(phone1.get_text())

    email_span = block.select_one("span.contactinfo_email")
    if email_span:
        email_text = clean_text(email_span.get_text(" ", strip=True))
        email_match = re.search(
            r'([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})',
            email_text,
            re.I)
        if email_match:
            result["Applicant Mail"] = email_match.group(1)

    if not result["Applicant Name"]:
        text = clean_text(block.get_text("\n", strip=True))
        lines = [clean_text(x) for x in text.split("\n") if clean_text(x)]
        if lines:
            result["Applicant Name"] = lines[0]
            if len(lines) > 1 and not re.search(r'\d', lines[1]):
                result["Applicant Company Name"] = lines[1]

    return result


def extract_licensed_professional(soup):
    result = {
        "Licensed Professional Name": "",
        "Licensed Professional Mail": "",
        "Licensed Professional Company Name": "",
        "Licensed Professional Address": "",
        "Licensed Professional Mailing City, State, Zip": "",
        "Licensed Professional Business Phone": "",
        "Licensed Professional Tax No": "",
        "Licensed Professional Contractor No": "",
    }

    table = soup.find("table", id="tbl_licensedps")
    if not table:
        block = find_label_block(soup, "Licensed Professional:")
        if not block:
            return result
        text = clean_text(block.get_text("\n", strip=True))
    else:
        text = clean_text(table.get_text("\n", strip=True))

    lines = [clean_text(x) for x in text.split("\n") if clean_text(x)]
    cleaned = [
        l for l in lines if l.lower() not in {
            "licensed professional:",
            "licensed professional"}]
    lines = cleaned

    if not lines:
        return result

    result["Licensed Professional Name"] = lines[0]

    city_pat = re.compile(
        r".+,\s*[A-Z]{2},\s*\d{5}(?:-\d{4})?$|.+,\s*[A-Z]{2},\s*[\dA-Z\- ]+$", re.I)
    city_idx = -1
    for i, line in enumerate(lines):
        if city_pat.match(line):
            city_idx = i
            break

    if city_idx != -1:
        result["Licensed Professional Mailing City, State, Zip"] = lines[city_idx]
        if city_idx > 0:
            result["Licensed Professional Address"] = lines[city_idx - 1]
        if city_idx > 1:
            result["Licensed Professional Company Name"] = lines[city_idx - 2]

    email_match = re.search(
        r'([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})', text, re.I)
    if email_match:
        result["Licensed Professional Mail"] = email_match.group(1)

    return result


def extract_project_description(soup):
    block = find_label_block(soup, "Project Description:")
    return clean_text(block.get_text(" ", strip=True)) if block else ""


def extract_owner(soup):
    result = {"Owner": "", "Owner Address": "",
              "Owner Mailing City, State, Zip": ""}
    block = find_label_block(soup, "Owner:")
    if not block:
        return result

    lines = [clean_text(x) for x in block.get_text(
        "\n", strip=True).split("\n") if clean_text(x)]
    lines = [l for l in lines if not re.fullmatch(r"[\*\-]+", l)]

    if len(lines) >= 1:
        result["Owner"] = lines[0].replace(" *", "").strip()
    if len(lines) >= 2:
        result["Owner Address"] = lines[1]
    if len(lines) >= 3:
        result["Owner Mailing City, State, Zip"] = lines[2]

    return result


def expand_more_detail_and_parcel(driver):
    """Expands optional detail sections using a short timeout for each."""
    # Use a short timeout since these buttons are optional and often missing
    short_wait = WebDriverWait(driver, 3)
    for btn_id in ["lnkMoreDetail", "lnkASI", "lnkParcelList"]:
        try:
            btn = short_wait.until(EC.element_to_be_clickable((By.ID, btn_id)))
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(2)
        except Exception:
            pass


def extract_parcel(soup):
    try:
        parcel_div = soup.find("div", id=re.compile(r"palParceList"))
        if not parcel_div:
            label = soup.find(
                "span", string=re.compile(
                    r"Parcel Number:", re.I))
            if label:
                val = label.find_next("span")
                if val:
                    return clean_text(val.text)

        if parcel_div:
            txt = clean_text(parcel_div.get_text(" ", strip=True))
            m = re.search(r'Parcel Number:\s*([A-Z0-9\-]+)', txt, re.I)
            if m:
                return m.group(1).strip()

        return ""
    except Exception:
        return ""


def parse_detail_page(driver, wait):
    expand_more_detail_and_parcel(driver)
    time.sleep(1.5)

    soup = BeautifulSoup(driver.page_source, "html.parser")

    data = {}
    data.update(extract_applicant(soup))
    data.update(extract_licensed_professional(soup))
    data["Project Description"] = extract_project_description(soup)
    data.update(extract_owner(soup))
    data["Parcel No"] = extract_parcel(soup)
    return data


def get_current_page_number(driver):
    try:
        ensure_results_frame(driver)
        elem = driver.find_element(
            By.XPATH,
            '//table[contains(@class,"aca_pagination")]//span[contains(@class,"SelectedPageButton")]'
        )
        return clean_text(elem.text)
    except Exception:
        return ""


def click_next_page(driver):
    try:
        ensure_results_frame(driver)
        current_page = get_current_page_number(driver)
        next_link = driver.find_element(
            By.XPATH,
            '//table[contains(@class,"aca_pagination")]//a[contains(normalize-space(.),"Next")]'
        )
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});",
            next_link)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", next_link)

        WebDriverWait(driver, 30).until(lambda d: get_current_page_number(
            d) != current_page and get_current_page_number(d) != "")
        wait_ready(driver, 30)
        time.sleep(1.5)
        return True
    except Exception:
        return False


def main():
    init_excel(OUTPUT_FILE)
    existing_record_numbers = get_existing_record_numbers(OUTPUT_FILE)
    s_no = get_next_serial_no(OUTPUT_FILE)
    driver = None

    try:
        log_message("Starting driver...")
        driver = create_driver()
        wait = WebDriverWait(driver, 30)

        log_message("Resolving search date range...")
        from_date, to_date = get_date_range(DATE_INPUT_FILE)

        log_message(f"Executing search for period: {from_date} - {to_date}")
        perform_automated_search(driver, wait, from_date, to_date)

        page_no = 1
        while True:
            log_message(f"Processing page: {page_no}")

            page_text = driver.page_source.lower()
            if "no results" in page_text or "no records found" in page_text:
                log_message("No records found for this date range.")
                break

            try:
                wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//*[@id="ctl00_PlaceHolderMain_dgvPermitList_gdvPermitList"]')))
            except Exception:
                if "no results" in page_text or "no records found" in page_text:
                    log_message("No records found.")
                    break
                else:
                    log_message("Results table not found.")
                    break

            time.sleep(2)
            row_count = get_all_data_rows_count(driver)
            log_message(f"Rows found on page {page_no}: {row_count}")

            if row_count == 0:
                break

            for row_pos in range(1, row_count + 1):
                row_output = {h: "" for h in HEADERS}
                row_output["S.No"] = s_no
                row_output["County"] = "Anaheim"
                row_output["State"] = "CA"

                try:
                    row = get_data_row_by_position(driver, row_pos)
                    summary = parse_summary_from_row(driver, row, wait)
                    row_output.update(summary)

                    current_record_number = clean_text(
                        row_output.get("Record Number", ""))
                    if current_record_number and current_record_number in existing_record_numbers:
                        log_message(f"Skipping: {current_record_number}")
                        continue

                    if is_record_clickable(row):
                        log_message(
                            f"Fetching details for: {current_record_number}")
                        open_record_in_new_tab(driver, row)
                        details = parse_detail_page(driver, wait)
                        row_output.update(details)
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                        wait_ready(driver, 30)
                        time.sleep(1)

                    append_row_to_excel(OUTPUT_FILE, row_output)
                    existing_record_numbers.add(current_record_number)
                    s_no += 1

                except Exception as e:
                    log_message(f"Error processing row {row_pos}: {e}")
                    if len(driver.window_handles) > 1:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])

            if not click_next_page(driver):
                break
            page_no += 1

    except Exception as e:
        log_message(f"Critical error: {e}")
    finally:
        if driver:
            driver.quit()
        log_message("Scraping completed.")


if __name__ == "__main__":
    main()
